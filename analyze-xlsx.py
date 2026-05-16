#!/usr/bin/env python3
"""分析 Excel 文件结构与性能瓶颈"""

import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
import json
from openpyxl import load_workbook

def analyze_xlsx_structure(file_path):
    """分析 xlsx 文件的内部结构"""
    file_path = Path(file_path)
    file_size_mb = file_path.stat().st_size / (1024 * 1024)
    
    print(f"📊 分析文件: {file_path.name}")
    print(f"📦 文件大小: {file_size_mb:.2f} MB\n")
    
    # 分析 ZIP 内部文件大小
    print("=" * 60)
    print("1️⃣  ZIP 内部文件大小分布")
    print("=" * 60)
    
    zip_analysis = {}
    total_uncompressed = 0
    
    with zipfile.ZipFile(file_path, 'r') as zf:
        for info in zf.infolist():
            uncompressed_size = info.file_size
            compressed_size = info.compress_size
            total_uncompressed += uncompressed_size
            
            category = "其他"
            if "sheet" in info.filename.lower():
                category = "Sheet XML"
            elif "drawing" in info.filename.lower() or "image" in info.filename.lower():
                category = "图片/绘图"
            elif "styles.xml" in info.filename:
                category = "样式表"
            elif "sharedStrings" in info.filename:
                category = "共享字符串"
            elif "theme" in info.filename:
                category = "主题"
            
            if category not in zip_analysis:
                zip_analysis[category] = {"files": [], "uncompressed": 0, "compressed": 0}
            
            zip_analysis[category]["files"].append({
                "name": info.filename,
                "uncompressed_kb": uncompressed_size / 1024,
                "compressed_kb": compressed_size / 1024
            })
            zip_analysis[category]["uncompressed"] += uncompressed_size
            zip_analysis[category]["compressed"] += compressed_size
    
    for category in sorted(zip_analysis.keys(), 
                          key=lambda x: zip_analysis[x]["uncompressed"], 
                          reverse=True):
        data = zip_analysis[category]
        print(f"\n📁 {category}:")
        print(f"   未压缩: {data['uncompressed']/1024:.1f} KB | 压缩: {data['compressed']/1024:.1f} KB | 比率: {100*data['compressed']/max(1,data['uncompressed']):.1f}%")
        
        if len(data["files"]) <= 5:
            for f in data["files"]:
                print(f"     └─ {f['name']}: {f['uncompressed_kb']:.1f} KB")
        else:
            for f in sorted(data["files"], key=lambda x: x["uncompressed_kb"], reverse=True)[:3]:
                print(f"     └─ {f['name']}: {f['uncompressed_kb']:.1f} KB")
            print(f"     └─ ... 还有 {len(data['files'])-3} 个文件")
    
    # 分析 Sheet 内容
    print("\n" + "=" * 60)
    print("2️⃣  Sheet 内容分析")
    print("=" * 60)
    
    wb = load_workbook(file_path, data_only=False)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n📄 Sheet: {sheet_name}")
        print(f"   行数 (rowCount): {ws.rowCount}")
        print(f"   列数 (colCount): {ws.max_column}")
        
        # 统计实际使用的行列
        actual_rows = 0
        actual_cols = 0
        cell_count = 0
        
        for row in ws.iter_rows():
            has_value = False
            for cell in row:
                if cell.value is not None or cell.fill != None or cell.font != None:
                    has_value = True
                    cell_count += 1
                    actual_cols = max(actual_cols, cell.column)
            if has_value:
                actual_rows = max(actual_rows, cell.row)
        
        print(f"   实际使用行数: {actual_rows}")
        print(f"   实际使用列数: {actual_cols}")
        print(f"   有内容的单元格数: {cell_count}")
        
        # 检查合并单元格
        if ws.merged_cells:
            print(f"   合并单元格数: {len(ws.merged_cells)}")
        
        # 检查图片/绘图
        if hasattr(ws, '_images') and ws._images:
            print(f"   ⚠️  包含 {len(ws._images)} 张图片")
        
        if hasattr(ws, '_drawing') and ws._drawing:
            print(f"   ⚠️  包含绘图对象")
    
    # 分析 Sheet XML 文件大小
    print("\n" + "=" * 60)
    print("3️⃣  Sheet XML 文件详细分析")
    print("=" * 60)
    
    with zipfile.ZipFile(file_path, 'r') as zf:
        for info in zf.infolist():
            if "sheet" in info.filename.lower() and info.filename.endswith(".xml"):
                size_kb = info.file_size / 1024
                print(f"\n📋 {info.filename}: {size_kb:.1f} KB")
                
                # 读取 XML 并分析元素
                try:
                    xml_content = zf.read(info.filename)
                    root = ET.fromstring(xml_content)
                    
                    # 统计关键元素
                    ns = {'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    
                    # 找行列
                    rows = root.findall('.//ss:row', ns) or root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row')
                    cells = root.findall('.//ss:c', ns) or root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c')
                    
                    print(f"   XML 元素 - 行: {len(rows)}, 单元格: {len(cells)}")
                    
                    # 查找特别大的单元格值
                    large_cells = []
                    for cell in cells[:100]:  # 仅检查前100个
                        v_elem = cell.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                        if v_elem is not None and v_elem.text and len(str(v_elem.text)) > 100:
                            large_cells.append(len(str(v_elem.text)))
                    
                    if large_cells:
                        print(f"   ⚠️  检测到大文本单元格: {len(large_cells)} 个 (最大 {max(large_cells)} 字符)")
                
                except Exception as e:
                    print(f"   ⚠️  XML 分析失败: {e}")
    
    # 分析共享字符串
    print("\n" + "=" * 60)
    print("4️⃣  共享字符串表 (sharedStrings.xml) 分析")
    print("=" * 60)
    
    with zipfile.ZipFile(file_path, 'r') as zf:
        if 'xl/sharedStrings.xml' in zf.namelist():
            shared_strings_size = zf.getinfo('xl/sharedStrings.xml').file_size
            print(f"\n📝 sharedStrings.xml 大小: {shared_strings_size / 1024:.1f} KB")
            
            try:
                xml_content = zf.read('xl/sharedStrings.xml')
                root = ET.fromstring(xml_content)
                
                si_elements = root.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si')
                print(f"   字符串条目数: {len(si_elements)}")
                
                # 找最长的字符串
                string_lengths = []
                for si in si_elements:
                    text_parts = si.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')
                    total_len = sum(len(t.text or "") for t in text_parts)
                    string_lengths.append(total_len)
                
                if string_lengths:
                    print(f"   最长字符串: {max(string_lengths)} 字符")
                    print(f"   平均长度: {sum(string_lengths) / len(string_lengths):.0f} 字符")
                    print(f"   超过1000字符的字符串: {len([x for x in string_lengths if x > 1000])} 个")
            
            except Exception as e:
                print(f"   ⚠️  分析失败: {e}")
    
    # 总结
    print("\n" + "=" * 60)
    print("🎯 性能瓶颈诊断")
    print("=" * 60)
    
    issues = []
    
    if file_size_mb > 5:
        issues.append(f"文件 > 5MB: {file_size_mb:.2f} MB")
    
    if 'Sheet XML' in zip_analysis and zip_analysis['Sheet XML']['uncompressed'] > 2 * 1024 * 1024:
        issues.append(f"Sheet XML 过大: {zip_analysis['Sheet XML']['uncompressed']/1024/1024:.1f} MB")
    
    if '图片/绘图' in zip_analysis and zip_analysis['图片/绘图']['uncompressed'] > 1 * 1024 * 1024:
        issues.append(f"图片/绘图过大: {zip_analysis['图片/绘图']['uncompressed']/1024/1024:.1f} MB")
    
    if 'Sheet XML' in zip_analysis and zip_analysis['Sheet XML']['compressed'] > 0:
        compression_ratio = zip_analysis['Sheet XML']['uncompressed'] / max(1, zip_analysis['Sheet XML']['compressed'])
        if compression_ratio > 20:
            issues.append(f"Sheet 压缩率极高 ({compression_ratio:.0f}:1)，可能有大量重复/空白")
    
    if issues:
        print("\n⚠️  发现的问题:")
        for issue in issues:
            print(f"   • {issue}")
    else:
        print("\n✅ 文件结构良好，未发现明显问题")

if __name__ == "__main__":
    target = Path("c:\\Users\\XXK\\Desktop\\work\\excel tools\\excel-tools\\test\\华锐捷.xlsx")
    if target.exists():
        analyze_xlsx_structure(target)
    else:
        print(f"❌ 文件不存在: {target}")
