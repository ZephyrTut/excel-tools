#!/usr/bin/env python3
"""优化 xlsx 文件：移除空白格式、图片等"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys
from datetime import datetime

input_file = Path(r"c:\Users\XXK\Desktop\work\excel tools\excel-tools\test\华锐捷.xlsx")
output_file = input_file.parent / f"{input_file.stem}_optimized.xlsx"

if not input_file.exists():
    print(f"❌ 输入文件不存在: {input_file}")
    sys.exit(1)

print(f"\n📦 优化 Excel 文件")
print(f"   输入: {input_file.name} ({input_file.stat().st_size / 1024 / 1024:.2f} MB)")

# 加载
print(f"\n⏳ 加载文件... (可能需要几秒)")
wb = load_workbook(input_file)

print(f"✅ 已加载 {len(wb.sheetnames)} 个 Sheet")

# 对每个 Sheet 优化
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n🔧 优化 {sheet_name}...")
    
    # 1. 找真实数据范围
    max_row = 0
    max_col = 0
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value is not None:
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    
    # 2. 删除超过实际范围的行
    if ws.max_row > max_row + 10:  # 保留 10 行缓冲
        rows_to_delete = ws.max_row - (max_row + 10)
        ws.delete_rows(max_row + 11, rows_to_delete)
        print(f"   删除空白行: {rows_to_delete}")
    
    # 3. 删除超过实际范围的列
    if ws.max_column > max_col + 5:  # 保留 5 列缓冲
        cols_to_delete = ws.max_column - (max_col + 5)
        ws.delete_cols(max_col + 6, cols_to_delete)
        print(f"   删除空白列: {cols_to_delete}")
    
    # 4. 移除图片（如果有）
    if hasattr(ws, '_images') and ws._images:
        image_count = len(ws._images)
        ws._images = []
        print(f"   移除图片: {image_count} 张")
    
    # 5. 移除绘图（如果有）
    if hasattr(ws, '_drawing') and ws._drawing:
        ws._drawing = None
        print(f"   移除绘图对象")
    
    # 6. 清理条件格式（可选）
    if hasattr(ws, 'conditional_formatting') and ws.conditional_formatting:
        print(f"   条件格式: {len(ws.conditional_formatting)} 个 (保留)")

# 保存
print(f"\n💾 保存优化后的文件...")
wb.save(output_file)

# 统计
orig_size = input_file.stat().st_size / (1024 * 1024)
new_size = output_file.stat().st_size / (1024 * 1024)
saved = orig_size - new_size
saved_pct = 100 * saved / orig_size if orig_size > 0 else 0

print(f"\n✅ 优化完成！")
print(f"   原始大小: {orig_size:.2f} MB")
print(f"   优化后: {new_size:.2f} MB")
print(f"   节省: {saved:.2f} MB ({saved_pct:.1f}%)")
print(f"\n📁 输出文件: {output_file.name}")

if saved_pct > 20:
    print(f"\n🎉 优化效果显著！节省了 {saved_pct:.0f}% 的空间")
    print(f"   建议用优化版本替换原文件")

print("\n")
