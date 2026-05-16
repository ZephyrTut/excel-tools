#!/usr/bin/env python3
"""检测并优化肥大 xlsx 文件"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys

xlsx_path = Path(r"c:\Users\XXK\Desktop\work\excel tools\excel-tools\test\华锐捷.xlsx")

if not xlsx_path.exists():
    print(f"❌ 文件不存在: {xlsx_path}")
    sys.exit(1)

print(f"\n📊 深度分析: {xlsx_path.name}\n")

# 加载工作簿
wb = load_workbook(xlsx_path, data_only=False)

# 按 Sheet 统计
total_rows = 0
total_cols = 0
total_cells_with_value = 0
max_row_used = {}
max_col_used = {}
merged_cell_count = 0
drawing_count = 0
style_count = len(wb.named_styles) if hasattr(wb, 'named_styles') else 0

print("=" * 70)
print("📄 Sheet 内容统计")
print("=" * 70)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    print(f"\n📋 {sheet_name}:")
    print(f"   定义的尺寸: {ws.rowCount} 行 × {ws.max_column} 列")
    
    # 找真实使用范围
    actual_rows = 0
    actual_cols = 0
    cells_used = 0
    
    for row in ws.iter_rows(min_row=1, max_row=ws.rowCount):
        for cell in row:
            if cell.value is not None:
                actual_rows = max(actual_rows, cell.row)
                actual_cols = max(actual_cols, cell.column)
                cells_used += 1
    
    max_row_used[sheet_name] = actual_rows or ws.rowCount
    max_col_used[sheet_name] = actual_cols or ws.max_column
    
    print(f"   实际数据范围: {actual_rows} 行 × {actual_cols} 列")
    print(f"   有值的单元格: {cells_used}")
    
    # 检查合并
    if ws.merged_cells:
        merge_count = len(ws.merged_cells)
        merged_cell_count += merge_count
        print(f"   合并单元格: {merge_count} 个")
    
    # 检查图画
    if hasattr(ws, '_images') and ws._images:
        print(f"   ⚠️  图片: {len(ws._images)} 张")
        drawing_count += len(ws._images)
    
    if hasattr(ws, '_drawing') and ws._drawing:
        print(f"   ⚠️  绘图对象存在")
    
    # 检查条件格式
    if hasattr(ws, 'conditional_formatting') and ws.conditional_formatting:
        print(f"   条件格式: {len(ws.conditional_formatting)} 个")
    
    total_rows += actual_rows or 1
    total_cols = max(total_cols, actual_cols or 1)
    total_cells_with_value += cells_used

print("\n" + "=" * 70)
print("🔍 根因诊断")
print("=" * 70)

issues = []

# 检查 1: 过度行列定义
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if ws.rowCount > max_row_used.get(sheet_name, 1) * 10:
        issues.append({
            "level": "🔴 严重",
            "issue": f"{sheet_name}: 行定义过多 ({ws.rowCount} 定义 vs {max_row_used[sheet_name]} 实际)",
            "impact": "可能有 1000+ 行的空白格式化",
            "fix": "删除未使用行"
        })

# 检查 2: 样式表
if style_count > 50:
    issues.append({
        "level": "🟡 中等",
        "issue": f"样式表过大: {style_count} 个样式",
        "impact": "样式定义占据额外 KB",
        "fix": "清理未使用样式"
    })

# 检查 3: 合并单元格
if merged_cell_count > 100:
    issues.append({
        "level": "🟡 中等",
        "issue": f"合并单元格过多: {merged_cell_count} 个",
        "impact": "增加 XML 大小",
        "fix": "评估是否必要"
    })

# 检查 4: 图片
if drawing_count > 5:
    issues.append({
        "level": "🔴 严重",
        "issue": f"图片/绘图过多: {drawing_count} 个",
        "impact": "占用大量文件空间",
        "fix": "删除不必要图片"
    })

if not issues:
    print("\n✅ 未发现明显问题")
else:
    print()
    for issue in issues:
        print(f"\n{issue['level']} {issue['issue']}")
        print(f"   影响: {issue['impact']}")
        print(f"   建议: {issue['fix']}")

print("\n" + "=" * 70)
print("📈 汇总统计")
print("=" * 70)
print(f"\nSheet 数: {len(wb.sheetnames)}")
print(f"总行数 (实际): {total_rows}")
print(f"总列数 (实际): {total_cols}")
print(f"有值单元格总数: {total_cells_with_value}")
print(f"合并单元格总数: {merged_cell_count}")
print(f"样式定义数: {style_count}")
print(f"图片/绘图数: {drawing_count}")

file_size_mb = xlsx_path.stat().st_size / (1024 * 1024)
avg_cell_size = (xlsx_path.stat().st_size / max(1, total_cells_with_value)) if total_cells_with_value > 0 else 0
print(f"\n文件大小: {file_size_mb:.2f} MB")
print(f"平均每个有值单元格: {avg_cell_size:.0f} 字节")

if avg_cell_size > 500:
    print(f"\n⚠️  警告: 平均每个单元格占用 {avg_cell_size:.0f} 字节（应 < 100）")
    print("   可能的原因:")
    print("   • 大量空白单元格被格式化")
    print("   • 图片/对象嵌入")
    print("   • 样式数据冗余")

print("\n")
